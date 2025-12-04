#!/usr/bin/env python3
"""
Automatisierte Qualitätstests für den FH-SWiFty-Chatbot.
Lädt Testfragen aus JSON, sendet sie an den Bot und exportiert Ergebnisse.

Verwendung:
    python testing_chatbot/test_auswertung.py
"""

import asyncio
import json
import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any
import time

# Projekt-Root zum Python-Path hinzufügen
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))
sys.path.insert(0, str(project_root / "fh-swifty-chatbot"))

from langchain_openai import ChatOpenAI
from langgraph.prebuilt import create_react_agent
from langchain_core.messages import HumanMessage

try:
    from helpers.tools import find_info_on_fhswf_website
    from helpers.prompts import prompt_langgraph
except ImportError as e:
    print(f"Import-Fehler: {e}")
    print("Bitte alle Abhängigkeiten installieren.")
    sys.exit(1)


class Colors:
    GREEN = '\033[92m'
    RED = '\033[91m'
    RESET = '\033[0m'


class ChatbotEvaluator:
    
    def __init__(self, test_data_path: str):
        self.test_data_path = Path(test_data_path)
        self.test_data: Dict[str, Any] = {}
        self.results: List[Dict[str, Any]] = []
        self.agent = None
        self.start_time = None
        self.end_time = None
        
    def load_test_data(self) -> bool:
        """Lädt die Testdaten aus der JSON-Datei"""
        try:
            with open(self.test_data_path, 'r', encoding='utf-8') as f:
                self.test_data = json.load(f)
            print(f"{Colors.GREEN}Testdaten erfolgreich geladen{Colors.RESET}")
            return True
        except FileNotFoundError:
            print(f"{Colors.RED}Fehler: Datei nicht gefunden - {self.test_data_path}{Colors.RESET}")
            return False
        except json.JSONDecodeError as e:
            print(f"{Colors.RED}JSON-Fehler: {e}{Colors.RESET}")
            return False
    
    def initialize_agent(self) -> bool:
        """Startet den Chatbot-Agent mit GPT-4o"""
        try:
            print("Initialisiere Chatbot-Agent...")
            model = ChatOpenAI(model="gpt-4o", streaming=False)
            tools = [find_info_on_fhswf_website]
            self.agent = create_react_agent(
                model=model, 
                tools=tools, 
                prompt=prompt_langgraph
            )
            print(f"{Colors.GREEN}Agent erfolgreich initialisiert{Colors.RESET}")
            return True
        except Exception as e:
            print(f"{Colors.RED}Fehler beim Starten: {e}{Colors.RESET}")
            return False
    
    async def ask_chatbot(self, question: str) -> Dict[str, Any]:
        """Sendet eine Frage an den Chatbot und misst die Antwortzeit"""
        if not self.agent:
            return {
                "answer": "",
                "response_time": 0,
                "error": "Agent nicht initialisiert"
            }
        
        start = time.time()
        try:
            response_parts = []
            # Frage an den Agent schicken
            async for msg, metadata in self.agent.astream(
                {"messages": [HumanMessage(question)]},
                stream_mode="messages",
            ):
                node = (metadata or {}).get("langgraph_node")
                content = (msg.content or "")
                
                # Nur Antworten vom Agent sammeln, keine Tool-Outputs
                if content and node != "tools":
                    response_parts.append(content)
            
            answer = "".join(response_parts).strip()
            response_time = time.time() - start
            
            return {
                "answer": answer,
                "response_time": round(response_time, 2),
                "error": None
            }
        except Exception as e:
            response_time = time.time() - start
            return {
                "answer": "",
                "response_time": round(response_time, 2),
                "error": str(e)
            }
    
    def extract_tests_from_data(self) -> List[Dict[str, Any]]:
        """Holt die Tests aus der JSON - jetzt schon in flacher Struktur"""
        # Die neue Struktur hat die Tests schon direkt als Liste
        tests = self.test_data.get("tests", [])
        
        if not tests:
            print(f"{Colors.RED}Warnung: Keine Tests in der Datei gefunden{Colors.RESET}")
        
        return tests
    
    async def run_evaluation(self):
        """Führt die gesamte Evaluation durch - Test für Test"""
        print("\n" + "="*70)
        print("  FH-SWiFty Chatbot - Qualitätsauswertung")
        print("="*70 + "\n")
        
        self.start_time = datetime.now()
        
        if not self.load_test_data():
            return False
        
        if not self.initialize_agent():
            return False
        
        all_tests = self.extract_tests_from_data()
        total_tests = len(all_tests)
        
        print(f"\nGefundene Tests: {total_tests}")
        print("Start der Evaluation...\n")
        
        for idx, test in enumerate(all_tests, 1):
            # Fortschritt anzeigen
            print(f"[{idx}/{total_tests}] {test.get('test_id')}: ", end="")
            frage = test.get('testfrage', '')
            print(f"{frage[:60]}{'...' if len(frage) > 60 else ''}")
            
            # Frage an den Bot senden
            response = await self.ask_chatbot(frage)
            
            # Ergebnis zusammenstellen (Test-Daten + neue Felder)
            result = {
                **test,
                "tatsaechliche_antwort": response['answer'],
                "antwortzeit_sekunden": response['response_time'],
                "fehler": response['error'],
                "timestamp": datetime.now().isoformat()
            }
            
            self.results.append(result)
            
            # Status ausgeben
            if response['error']:
                print(f"  {Colors.RED}Fehler: {response['error']}{Colors.RESET}")
            else:
                print(f"  {Colors.GREEN}OK ({response['response_time']}s){Colors.RESET}")
            
            # Kleine Pause zwischen den Tests (wegen API Rate Limits)
            if idx < total_tests:
                await asyncio.sleep(1)
        
        self.end_time = datetime.now()
        
        print(f"\n{Colors.GREEN}Evaluation abgeschlossen!{Colors.RESET}")
        print(f"Gesamtdauer: {(self.end_time - self.start_time).total_seconds():.2f} Sekunden\n")
        
        return True
    
    def save_results(self, output_dir: str = "testing_chatbot/test_results"):
        """Speichert die Ergebnisse als JSON und Excel"""
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # JSON speichern
        json_file = output_path / f"chatbot_evaluation_{timestamp}.json"
        json_data = {
            "metadata": {
                "test_version": self.test_data.get("metadata", {}).get("test_version", "1.0"),
                "institution": self.test_data.get("metadata", {}).get("institution", "FH Südwestfalen"),
                "evaluation_date": self.start_time.isoformat() if self.start_time else None,
                "total_tests": len(self.results),
                "duration_seconds": (self.end_time - self.start_time).total_seconds() if self.start_time and self.end_time else 0
            },
            "results": self.results,
            "summary": self._generate_summary()
        }
        
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        
        print(f"{Colors.GREEN}JSON gespeichert: {json_file}{Colors.RESET}")
        
        # Excel speichern (falls openpyxl installiert ist)
        try:
            self._save_to_excel(output_path, timestamp)
        except ImportError:
            print("Warnung: openpyxl nicht installiert. Excel-Export übersprungen.")
            print("Installieren mit: uv pip install openpyxl")
    
    def _save_to_excel(self, output_path: Path, timestamp: str):
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise
        
        excel_file = output_path / f"chatbot_evaluation_{timestamp}.xlsx"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Evaluation Results"
        
        headers = [
            "Test ID", "Kategorie", "Test Typ", "Szenario/Subkategorie",
            "Testfrage", "Erwartete Antwort", "Tatsächliche Antwort",
            "Antwortzeit (s)"
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        for row_idx, result in enumerate(self.results, 2):
            ws.cell(row=row_idx, column=1, value=result.get('test_id'))
            ws.cell(row=row_idx, column=2, value=result.get('kategorie'))
            ws.cell(row=row_idx, column=3, value=result.get('test_typ'))
            ws.cell(row=row_idx, column=4, value=result.get('szenario') or result.get('subkategorie', ''))
            ws.cell(row=row_idx, column=5, value=result.get('testfrage'))
            ws.cell(row=row_idx, column=6, value=result.get('erwartete_antwort'))
            ws.cell(row=row_idx, column=7, value=result.get('tatsaechliche_antwort'))
            ws.cell(row=row_idx, column=8, value=result.get('antwortzeit_sekunden'))
        
        column_widths = {
            1: 15,  # Test ID
            2: 25,  # Kategorie
            3: 15,  # Test Typ
            4: 20,  # Szenario
            5: 50,  # Testfrage
            6: 50,  # Erwartete Antwort
            7: 60,  # Tatsächliche Antwort
            8: 15   # Antwortzeit
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        
        ws_summary = wb.create_sheet(title="Zusammenfassung")
        summary = self._generate_summary()
        
        row = 1
        ws_summary.cell(row=row, column=1, value="Evaluations-Zusammenfassung")
        row += 2
        
        for key, value in summary.items():
            ws_summary.cell(row=row, column=1, value=key)
            ws_summary.cell(row=row, column=2, value=str(value))
            row += 1
        
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 40
        
        wb.save(excel_file)
        print(f"{Colors.GREEN}Excel gespeichert: {excel_file}{Colors.RESET}")
    
    def _generate_summary(self) -> Dict[str, Any]:
        total = len(self.results)
        errors = sum(1 for r in self.results if r.get('fehler'))
        successful = total - errors
        
        response_times = [r.get('antwortzeit_sekunden', 0) for r in self.results if not r.get('fehler')]
        avg_response_time = sum(response_times) / len(response_times) if response_times else 0
        
        kategorien = {}
        for result in self.results:
            kat = result.get('kategorie', 'Unbekannt')
            kategorien[kat] = kategorien.get(kat, 0) + 1
        
        return {
            "Gesamt Tests": total,
            "Erfolgreich": successful,
            "Fehler": errors,
            "Erfolgsrate (%)": round((successful / total * 100), 2) if total > 0 else 0,
            "Durchschnittliche Antwortzeit (s)": round(avg_response_time, 2),
            "Min Antwortzeit (s)": round(min(response_times), 2) if response_times else 0,
            "Max Antwortzeit (s)": round(max(response_times), 2) if response_times else 0,
            "Tests nach Kategorie": kategorien,
            "Start Zeit": self.start_time.isoformat() if self.start_time else None,
            "End Zeit": self.end_time.isoformat() if self.end_time else None,
            "Dauer (s)": round((self.end_time - self.start_time).total_seconds(), 2) if self.start_time and self.end_time else 0
        }
    
    def print_summary(self):
        summary = self._generate_summary()
        
        print("\n" + "="*70)
        print("  Zusammenfassung")
        print("="*70 + "\n")
        
        print(f"Gesamt Tests: {summary['Gesamt Tests']}")
        print(f"{Colors.GREEN}Erfolgreich: {summary['Erfolgreich']}{Colors.RESET}")
        print(f"{Colors.RED}Fehler: {summary['Fehler']}{Colors.RESET}")
        print(f"Erfolgsrate: {summary['Erfolgsrate (%)']}%")
        print(f"Durchschn. Antwortzeit: {summary['Durchschnittliche Antwortzeit (s)']}s")
        print(f"Min/Max Antwortzeit: {summary['Min Antwortzeit (s)']}s / {summary['Max Antwortzeit (s)']}s")
        
        print("\nTests nach Kategorie:")
        for kat, count in summary['Tests nach Kategorie'].items():
            print(f"  - {kat}: {count}")
        
        print(f"\nDauer: {summary['Dauer (s)']}s")
        print("="*70 + "\n")


async def main():
    # Pfad zur Testdaten-Datei (neue vereinfachte Struktur)
    # Für Quick-Test: chatbot_tests_mini.json (nur 3 Tests)
    # Für volle Tests: chatbot_tests_simple.json (alle 88 Tests)
    test_data_file = "testing_chatbot/test_data/chatbot_tests_simple.json"
    
    evaluator = ChatbotEvaluator(test_data_file)
    
    # Tests durchführen
    success = await evaluator.run_evaluation()
    
    if success:
        # Ergebnisse speichern (JSON + Excel)
        evaluator.save_results()
        evaluator.print_summary()
        
        print(f"{Colors.GREEN}Evaluation erfolgreich abgeschlossen!{Colors.RESET}\n")
        return 0
    else:
        print(f"{Colors.RED}Evaluation fehlgeschlagen{Colors.RESET}\n")
        return 1


if __name__ == "__main__":
    exit_code = asyncio.run(main())
    sys.exit(exit_code)

